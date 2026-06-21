from fastapi import FastAPI, HTTPException, Depends, Header, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, model_validator
from typing import List, Optional
from datetime import datetime
from email.message import EmailMessage
from email.utils import formataddr
from html import escape
from io import BytesIO
from pathlib import Path
from urllib.parse import quote_plus
from sqlalchemy import Boolean, create_engine, inspect, Column, Integer, String, DateTime
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import hashlib
import logging
import os
import secrets
import smtplib
import string
import time

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR.parent / ".env")
load_dotenv(BASE_DIR / ".env", override=False)

# Database Setup
DEFAULT_DATABASE_URL = "postgresql+psycopg2://feedrh:feedrh@localhost:5432/feedrh"

def normalize_database_url(database_url: str) -> str:
    if database_url.startswith("postgres://"):
        return database_url.replace(
            "postgres://",
            "postgresql://",
            1,
        )
    return database_url

def get_env_value(*names: str) -> Optional[str]:
    for name in names:
        value = os.getenv(name)
        if value:
            return value
    return None

def build_database_url_from_parts() -> Optional[str]:
    host = get_env_value("POSTGRES_HOST", "POSTGRESQL_HOST", "DB_HOST")
    database = get_env_value("POSTGRES_DB", "POSTGRES_DATABASE", "POSTGRESQL_DATABASE", "DB_NAME")
    user = get_env_value("POSTGRES_USER", "POSTGRES_USERNAME", "POSTGRESQL_USER", "DB_USER")
    password = get_env_value("POSTGRES_PASSWORD", "POSTGRESQL_PASSWORD", "DB_PASSWORD")
    port = get_env_value("POSTGRES_PORT", "POSTGRESQL_PORT", "DB_PORT") or "5432"

    if not all([host, database, user, password]):
        return None

    return (
        f"postgresql+psycopg2://{quote_plus(user)}:{quote_plus(password)}"
        f"@{host}:{port}/{quote_plus(database)}"
    )

def get_database_url() -> str:
    database_url = (
        os.getenv("DATABASE_URL")
        or os.getenv("SQLALCHEMY_DATABASE_URL")
        or build_database_url_from_parts()
        or os.getenv("DB_URL")
        or DEFAULT_DATABASE_URL
    )
    return normalize_database_url(database_url)

SQLALCHEMY_DATABASE_URL = get_database_url()

connect_args = {}
if SQLALCHEMY_DATABASE_URL.startswith("sqlite"):
    connect_args["check_same_thread"] = False

engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args=connect_args, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()
logger = logging.getLogger("feedrh")
logger.setLevel(logging.INFO)

ETAPAS_FUNIL = {
    1: "Fila de Espera",
    2: "Divulgação",
    3: "Triagem",
    4: "Entrevista Inicial",
    5: "Testes Psicológicos",
    6: "Parecer Psicológico",
    7: "Entrevista com Gestor",
    8: "Aguardando Retorno",
    9: "Finalizada",
}

def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode()).hexdigest()

# SQLAlchemy Models
class UserModel(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    nome = Column(String, index=True)
    email = Column(String, unique=True, index=True)
    empresa = Column(String)
    perfil = Column(String)  # "RH" ou "GESTOR"
    senha_hash = Column(String)
    must_change_password = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow)
    ultimo_reset_senha = Column(DateTime, nullable=True)

class EmpresaModel(Base):
    __tablename__ = "empresas"
    id = Column(Integer, primary_key=True, index=True)
    nome = Column(String, unique=True, index=True)

class VagaModel(Base):
    __tablename__ = "vagas"
    id = Column(Integer, primary_key=True, index=True)
    cargo = Column(String)
    data_abertura = Column(DateTime, default=datetime.utcnow)
    empresa_destinada = Column(String)
    senioridade = Column(String)
    resumo_requisitos = Column(String, nullable=True)
    requisitos_obrigatorios = Column(String, nullable=True)
    tipo = Column(String)  # "Nova posição" ou "Substituição"
    profissional_substituido = Column(String, nullable=True)
    justificativa_substituicao = Column(String, nullable=True)
    solicitante_id = Column(Integer)
    status_decisao_diretoria = Column(String, default="Pendente")
    justificativa_negativa = Column(String, nullable=True)
    quantidade_congelamentos = Column(Integer, default=0)
    etapa_funil = Column(Integer, default=1)
    data_finalizacao = Column(DateTime, nullable=True)

class VagaHistoricoModel(Base):
    __tablename__ = "vagas_historico"
    id = Column(Integer, primary_key=True, index=True)
    vaga_id = Column(Integer, index=True)
    data_registro = Column(DateTime, default=datetime.utcnow)
    usuario_id = Column(Integer)
    usuario_nome = Column(String)
    acao = Column(String)
    status_anterior = Column(String, nullable=True)
    status_novo = Column(String, nullable=True)
    justificativa = Column(String, nullable=True)

def garantir_colunas_vagas():
    colunas = {coluna["name"] for coluna in inspect(engine).get_columns("vagas")}

    with engine.begin() as connection:
        if "quantidade_congelamentos" not in colunas:
            connection.exec_driver_sql("ALTER TABLE vagas ADD COLUMN quantidade_congelamentos INTEGER DEFAULT 0")
        if "resumo_requisitos" not in colunas:
            connection.exec_driver_sql("ALTER TABLE vagas ADD COLUMN resumo_requisitos TEXT DEFAULT ''")
        if "requisitos_obrigatorios" not in colunas:
            connection.exec_driver_sql("ALTER TABLE vagas ADD COLUMN requisitos_obrigatorios TEXT DEFAULT ''")
        if "justificativa_negativa" not in colunas:
            connection.exec_driver_sql("ALTER TABLE vagas ADD COLUMN justificativa_negativa TEXT")
        connection.exec_driver_sql(
            "UPDATE vagas SET quantidade_congelamentos = 1 "
            "WHERE status_decisao_diretoria = 'Congelada' "
            "AND (quantidade_congelamentos IS NULL OR quantidade_congelamentos = 0)"
        )
        connection.exec_driver_sql("UPDATE vagas SET resumo_requisitos = '' WHERE resumo_requisitos IS NULL")
        connection.exec_driver_sql("UPDATE vagas SET requisitos_obrigatorios = '' WHERE requisitos_obrigatorios IS NULL")

def garantir_colunas_users():
    colunas = {coluna["name"] for coluna in inspect(engine).get_columns("users")}

    with engine.begin() as connection:
        if "must_change_password" not in colunas:
            connection.exec_driver_sql("ALTER TABLE users ADD COLUMN must_change_password BOOLEAN DEFAULT FALSE")
        if "created_at" not in colunas:
            connection.exec_driver_sql("ALTER TABLE users ADD COLUMN created_at TIMESTAMP")
        if "updated_at" not in colunas:
            connection.exec_driver_sql("ALTER TABLE users ADD COLUMN updated_at TIMESTAMP")
        if "ultimo_reset_senha" not in colunas:
            connection.exec_driver_sql("ALTER TABLE users ADD COLUMN ultimo_reset_senha TIMESTAMP")
        connection.exec_driver_sql("UPDATE users SET must_change_password = FALSE WHERE must_change_password IS NULL")
        connection.exec_driver_sql("UPDATE users SET created_at = CURRENT_TIMESTAMP WHERE created_at IS NULL")
        connection.exec_driver_sql("UPDATE users SET updated_at = CURRENT_TIMESTAMP WHERE updated_at IS NULL")

def get_database_retry_config() -> tuple[int, float]:
    try:
        attempts = int(os.getenv("DB_CONNECT_RETRIES", "10"))
    except ValueError:
        attempts = 10

    try:
        delay = float(os.getenv("DB_CONNECT_RETRY_SECONDS", "2"))
    except ValueError:
        delay = 2.0

    return max(attempts, 1), max(delay, 0.0)

def initialize_database() -> None:
    attempts, delay = get_database_retry_config()

    for attempt in range(1, attempts + 1):
        try:
            Base.metadata.create_all(bind=engine)
            garantir_colunas_vagas()
            garantir_colunas_users()
            return
        except Exception:
            if attempt == attempts:
                logger.exception("Nao foi possivel conectar ao banco de dados apos %s tentativa(s).", attempts)
                raise

            logger.warning(
                "Banco de dados indisponivel na tentativa %s/%s. Nova tentativa em %.1fs.",
                attempt,
                attempts,
                delay,
                exc_info=True,
            )
            time.sleep(delay)

# Pydantic Schemas
class LoginRequest(BaseModel):
    email: str
    senha: str

class LoginResponse(BaseModel):
    id: int
    nome: str
    email: str
    empresa: str
    perfil: str

class UserCreate(BaseModel):
    nome: str
    email: str
    empresa: str
    perfil: str  # "RH" ou "GESTOR"
    senha: str

class UserUpdate(BaseModel):
    nome: str
    email: str
    empresa: str
    perfil: str

class UserResponse(BaseModel):
    id: int
    nome: str
    email: str
    empresa: str
    perfil: str
    must_change_password: bool = False
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    ultimo_reset_senha: Optional[datetime] = None

    class Config:
        from_attributes = True

class UserCreateResponse(UserResponse):
    message: str
    email_enviado: Optional[bool] = None

class ResetPasswordResponse(BaseModel):
    message: str
    email_enviado: bool

class EmpresaCreate(BaseModel):
    nome: str

class EmpresaUpdate(BaseModel):
    nome: str

class EmpresaResponse(BaseModel):
    id: int
    nome: str

    class Config:
        from_attributes = True

class VagaHistoricoResponse(BaseModel):
    id: int
    data_registro: datetime
    usuario_id: int
    usuario_nome: str
    acao: str
    status_anterior: Optional[str] = None
    status_novo: Optional[str] = None
    justificativa: Optional[str] = None

    class Config:
        from_attributes = True

class VagaCreate(BaseModel):
    cargo: str
    empresa_destinada: str
    senioridade: str
    resumo_requisitos: str
    requisitos_obrigatorios: str
    tipo: str
    profissional_substituido: Optional[str] = None
    justificativa_substituicao: Optional[str] = None

    @model_validator(mode='after')
    def check_dados_obrigatorios(self) -> 'VagaCreate':
        if not self.resumo_requisitos or not self.resumo_requisitos.strip():
            raise ValueError("O resumo dos requisitos é obrigatório.")
        if not self.requisitos_obrigatorios or not self.requisitos_obrigatorios.strip():
            raise ValueError("Os requisitos obrigatórios devem ser informados.")
        if self.tipo == "Substituição":
            if not self.profissional_substituido or not self.profissional_substituido.strip():
                raise ValueError("O profissional substituído é obrigatório para vagas de Substituição.")
            if not self.justificativa_substituicao or not self.justificativa_substituicao.strip():
                raise ValueError("A justificativa é obrigatória para vagas de Substituição.")
        return self

class VagaResponse(BaseModel):
    id: int
    cargo: str
    data_abertura: datetime
    empresa_destinada: str
    senioridade: str
    resumo_requisitos: Optional[str] = None
    requisitos_obrigatorios: Optional[str] = None
    tipo: str
    profissional_substituido: Optional[str] = None
    justificativa_substituicao: Optional[str] = None
    solicitante_id: int
    solicitante_nome: Optional[str] = None
    solicitante_email: Optional[str] = None
    solicitante_empresa: Optional[str] = None
    status_decisao_diretoria: str
    justificativa_negativa: Optional[str] = None
    quantidade_congelamentos: int = 0
    etapa_funil: int
    data_finalizacao: Optional[datetime] = None
    posicao_fila_rh: Optional[int] = None
    historico: List[VagaHistoricoResponse] = Field(default_factory=list)

    class Config:
        from_attributes = True

class DecisaoDiretoriaUpdate(BaseModel):
    status: str
    justificativa_negativa: Optional[str] = None

class EtapaFunilUpdate(BaseModel):
    etapa: int

# FastAPI App
app = FastAPI(title="FeedRh API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"status": "ok", "service": "FeedRh API"}

@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(status_code=204)

@app.get("/health")
def health():
    return {"status": "ok"}

# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Auth Dependency via header X-User-Id
def get_current_user(x_user_id: str = Header(...), db: Session = Depends(get_db)):
    try:
        user_id = int(x_user_id)
    except ValueError:
        raise HTTPException(status_code=401, detail="X-User-Id header deve ser um número inteiro")
    user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return user

def require_rh(current_user: UserModel = Depends(get_current_user)):
    if current_user.perfil != "RH":
        raise HTTPException(status_code=403, detail="Acesso negado. Apenas usuários do RH possuem acesso a esta ação.")
    return current_user

def parse_data_filtro(valor: Optional[str], campo: str) -> Optional[datetime]:
    if not valor:
        return None
    try:
        return datetime.strptime(valor, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{campo} deve estar no formato YYYY-MM-DD")

def aplicar_filtros_vagas(
    query,
    gestor_id: Optional[int],
    data_inicio: Optional[str],
    data_fim: Optional[str],
    empresa: Optional[str] = None,
    senioridade: Optional[str] = None,
    etapa_funil: Optional[int] = None,
    status_decisao: Optional[str] = None,
):
    data_inicio_dt = parse_data_filtro(data_inicio, "data_inicio")
    data_fim_dt = parse_data_filtro(data_fim, "data_fim")

    if gestor_id is not None:
        query = query.filter(VagaModel.solicitante_id == gestor_id)
    if empresa:
        query = query.filter(VagaModel.empresa_destinada == normalizar_nome_empresa(empresa))
    if senioridade:
        query = query.filter(VagaModel.senioridade == senioridade)
    if etapa_funil is not None:
        if etapa_funil < 1 or etapa_funil > 9:
            raise HTTPException(status_code=400, detail="etapa_funil deve ser de 1 a 9")
        query = query.filter(VagaModel.etapa_funil == etapa_funil)
    if status_decisao:
        if status_decisao not in ["Pendente", "Aprovada", "Congelada", "Negada"]:
            raise HTTPException(status_code=400, detail="status_decisao inválido")
        query = query.filter(VagaModel.status_decisao_diretoria == status_decisao)
    if data_inicio_dt:
        query = query.filter(VagaModel.data_abertura >= data_inicio_dt)
    if data_fim_dt:
        query = query.filter(VagaModel.data_abertura <= data_fim_dt.replace(hour=23, minute=59, second=59, microsecond=999999))
    return query

def preencher_dados_vagas(db: Session, vagas: List[VagaModel]) -> List[VagaModel]:
    fila_rh = (
        db.query(VagaModel)
        .filter(VagaModel.status_decisao_diretoria == "Pendente")
        .order_by(VagaModel.data_abertura.asc(), VagaModel.id.asc())
        .all()
    )
    posicoes = {vaga.id: indice for indice, vaga in enumerate(fila_rh, start=1)}

    solicitante_ids = {vaga.solicitante_id for vaga in vagas}
    usuarios = {}
    if solicitante_ids:
        usuarios = {
            usuario.id: usuario
            for usuario in db.query(UserModel).filter(UserModel.id.in_(solicitante_ids)).all()
        }

    vaga_ids = [vaga.id for vaga in vagas]
    historico_por_vaga = {vaga_id: [] for vaga_id in vaga_ids}
    if vaga_ids:
        registros = (
            db.query(VagaHistoricoModel)
            .filter(VagaHistoricoModel.vaga_id.in_(vaga_ids))
            .order_by(VagaHistoricoModel.data_registro.asc(), VagaHistoricoModel.id.asc())
            .all()
        )
        for registro in registros:
            historico_por_vaga.setdefault(registro.vaga_id, []).append(registro)

    for vaga in vagas:
        solicitante = usuarios.get(vaga.solicitante_id)
        vaga.solicitante_nome = solicitante.nome if solicitante else "Usuário removido"
        vaga.solicitante_email = solicitante.email if solicitante else None
        vaga.solicitante_empresa = solicitante.empresa if solicitante else None
        vaga.posicao_fila_rh = posicoes.get(vaga.id)
        vaga.historico = historico_por_vaga.get(vaga.id, [])
    return vagas

def normalizar_nome_empresa(nome: str) -> str:
    return " ".join(nome.strip().split())

def validar_nome_empresa(nome: str) -> str:
    nome_normalizado = normalizar_nome_empresa(nome)
    if len(nome_normalizado) < 2:
        raise HTTPException(status_code=400, detail="Informe um nome de empresa válido")
    return nome_normalizado

def buscar_empresa_por_nome(db: Session, nome: str):
    return db.query(EmpresaModel).filter(EmpresaModel.nome.ilike(nome)).first()

def validar_dados_usuario(db: Session, nome: str, email: str, empresa: str, perfil: str, user_id_ignorar: Optional[int] = None):
    nome_normalizado = " ".join(nome.strip().split())
    email_normalizado = email.strip().lower()
    empresa_normalizada = normalizar_nome_empresa(empresa)

    if len(nome_normalizado) < 2:
        raise HTTPException(status_code=400, detail="Informe um nome de usuário válido")
    if "@" not in email_normalizado:
        raise HTTPException(status_code=400, detail="Informe um e-mail válido")
    if perfil not in ["RH", "GESTOR"]:
        raise HTTPException(status_code=400, detail="Perfil deve ser 'RH' ou 'GESTOR'")
    if not buscar_empresa_por_nome(db, empresa_normalizada):
        raise HTTPException(status_code=400, detail="Empresa não cadastrada")

    email_query = db.query(UserModel).filter(UserModel.email.ilike(email_normalizado))
    if user_id_ignorar is not None:
        email_query = email_query.filter(UserModel.id != user_id_ignorar)
    if email_query.first():
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")

    return nome_normalizado, email_normalizado, empresa_normalizada, perfil

def criar_empresa_se_nao_existir(db: Session, nome: str):
    nome_normalizado = normalizar_nome_empresa(nome)
    empresa = buscar_empresa_por_nome(db, nome_normalizado)
    if empresa:
        return empresa
    empresa = EmpresaModel(nome=nome_normalizado)
    db.add(empresa)
    db.commit()
    db.refresh(empresa)
    return empresa

def enviar_email_notificacao(destinatario: str, assunto: str, corpo: str, corpo_html: Optional[str] = None) -> bool:
    mail_host = os.getenv("MAIL_HOST")
    try:
        mail_port = int(os.getenv("MAIL_PORT", "587"))
    except ValueError:
        logger.warning("MAIL_PORT inválida. Usando porta 587.")
        mail_port = 587
    mail_user = os.getenv("MAIL_USER")
    mail_password = os.getenv("MAIL_PASSWORD")
    mail_from = os.getenv("MAIL_FROM") or mail_user
    mail_from_name = os.getenv("MAIL_FROM_NAME", "Sistema de Recrutamento")
    mail_use_ssl = os.getenv("MAIL_USE_SSL", "true" if mail_port == 465 else "false").lower() not in ["0", "false", "no", "nao"]
    mail_use_tls = (
        os.getenv("MAIL_USE_TLS", "false" if mail_use_ssl else "true").lower()
        not in ["0", "false", "no", "nao"]
    )

    if not mail_host or not mail_from:
        logger.error(
            "Envio de e-mail não configurado. Verifique MAIL_HOST e MAIL_FROM no .env. "
            "Destinatário: %s | Assunto: %s",
            destinatario,
            assunto,
        )
        return False

    mensagem = EmailMessage()
    mensagem["From"] = formataddr((mail_from_name, mail_from)) if mail_from_name else mail_from
    mensagem["To"] = destinatario
    mensagem["Subject"] = assunto
    mensagem.set_content(corpo)
    if corpo_html:
        mensagem.add_alternative(corpo_html, subtype="html")

    try:
        smtp_client = smtplib.SMTP_SSL if mail_use_ssl else smtplib.SMTP
        with smtp_client(mail_host, mail_port, timeout=10) as server:
            if mail_use_tls and not mail_use_ssl:
                server.starttls()
            if mail_user and mail_password:
                server.login(mail_user, mail_password)
            server.send_message(mensagem)
        logger.info("E-mail de notificação enviado para %s", destinatario)
        return True
    except Exception:
        logger.error("Não foi possível enviar e-mail de notificação para %s", destinatario, exc_info=True)
        return False

def montar_email_avanco_html(vaga: VagaModel, gestor: UserModel, resumo: str) -> str:
    etapa_nome = ETAPAS_FUNIL.get(vaga.etapa_funil, f"Etapa {vaga.etapa_funil}")
    progresso = max(0, min(100, round(((vaga.etapa_funil or 1) / 9) * 100)))
    app_url = (os.getenv("APP_URL") or os.getenv("FRONTEND_URL") or "http://localhost:4200").rstrip("/")
    dashboard_url = f"{app_url}/dashboard"

    cargo = escape(vaga.cargo or "Vaga")
    gestor_nome = escape(gestor.nome or "")
    resumo_html = escape(resumo).replace("\n", "<br>")
    empresa = escape(vaga.empresa_destinada or "Não informado")
    senioridade = escape(vaga.senioridade or "Não informado")
    status = escape(vaga.status_decisao_diretoria or "Não informado")
    etapa = escape(etapa_nome)
    tipo = escape(vaga.tipo or "Não informado")
    data_abertura = vaga.data_abertura.strftime("%d/%m/%Y") if vaga.data_abertura else "Não informado"

    return f"""<!doctype html>
<html lang="pt-BR">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
      @media only screen and (max-width: 640px) {{
        .container {{ width: 100% !important; }}
        .content {{ padding: 24px 18px !important; }}
        .hero {{ padding: 28px 22px !important; }}
        .title {{ font-size: 26px !important; line-height: 32px !important; }}
        .detail-cell {{ display: block !important; width: 100% !important; padding-right: 0 !important; padding-bottom: 12px !important; }}
        .button {{ display: block !important; }}
      }}
    </style>
  </head>
  <body style="margin:0; padding:0; background:#F3E5F5; font-family:Inter, Arial, Helvetica, sans-serif; color:#1F2937;">
    <div style="display:none; max-height:0; overflow:hidden; opacity:0;">Houve um novo avanço na vaga {cargo}.</div>
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F3E5F5; padding:32px 12px;">
      <tr>
        <td align="center">
          <table role="presentation" class="container" width="640" cellpadding="0" cellspacing="0" style="width:640px; max-width:640px; background:#FFFFFF; border-radius:20px; overflow:hidden; box-shadow:0 18px 45px rgba(49,27,146,0.14); border:1px solid #E9D5FF;">
            <tr>
              <td class="hero" style="padding:34px 34px 30px; background:#6200EE;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                  <tr>
                    <td style="font-size:13px; line-height:18px; color:#EDE7FF; font-weight:700; letter-spacing:0.08em; text-transform:uppercase;">
                      FeedRH
                    </td>
                    <td align="right">
                      <span style="display:inline-block; background:#B388FF; color:#311B92; border-radius:999px; padding:8px 12px; font-size:12px; font-weight:800;">
                        Novo avanço
                      </span>
                    </td>
                  </tr>
                </table>
                <h1 class="title" style="margin:22px 0 10px; color:#FFFFFF; font-size:32px; line-height:38px; font-weight:800; letter-spacing:0;">
                  Avanço na vaga:<br>{cargo}
                </h1>
                <p style="margin:0; color:#EDE7FF; font-size:15px; line-height:24px;">
                  Olá{", " + gestor_nome if gestor_nome else ""}. A vaga teve uma atualização no processo seletivo.
                </p>
              </td>
            </tr>
            <tr>
              <td class="content" style="padding:30px 34px 34px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:22px;">
                  <tr>
                    <td style="background:#FAF7FF; border:1px solid #E9D5FF; border-radius:16px; padding:20px;">
                      <p style="margin:0 0 8px; color:#6200EE; font-size:12px; line-height:16px; font-weight:800; letter-spacing:0.06em; text-transform:uppercase;">
                        Resumo do ocorrido
                      </p>
                      <p style="margin:0; color:#1F2937; font-size:16px; line-height:25px; font-weight:600;">
                        {resumo_html}
                      </p>
                    </td>
                  </tr>
                </table>

                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:22px;">
                  <tr>
                    <td style="padding-bottom:10px;">
                      <p style="margin:0; color:#311B92; font-size:15px; line-height:20px; font-weight:800;">Etapa atual</p>
                    </td>
                    <td align="right" style="padding-bottom:10px;">
                      <p style="margin:0; color:#6200EE; font-size:14px; line-height:20px; font-weight:800;">{vaga.etapa_funil or 1} / 9</p>
                    </td>
                  </tr>
                  <tr>
                    <td colspan="2" style="background:#E5E7EB; border-radius:999px; height:12px; overflow:hidden;">
                      <div style="background:#6200EE; width:{progresso}%; height:12px; border-radius:999px;"></div>
                    </td>
                  </tr>
                  <tr>
                    <td colspan="2" style="padding-top:8px;">
                      <p style="margin:0; color:#6B7280; font-size:13px; line-height:18px;">{progresso}% - {etapa}</p>
                    </td>
                  </tr>
                </table>

                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:24px;">
                  <tr>
                    <td class="detail-cell" width="50%" style="width:50%; padding-right:8px; padding-bottom:12px;">
                      <div style="border:1px solid #EEF2F7; border-radius:14px; padding:14px; background:#FFFFFF;">
                        <p style="margin:0 0 4px; color:#9CA3AF; font-size:11px; line-height:15px; font-weight:800; text-transform:uppercase;">Empresa</p>
                        <p style="margin:0; color:#111827; font-size:14px; line-height:20px; font-weight:700;">{empresa}</p>
                      </div>
                    </td>
                    <td class="detail-cell" width="50%" style="width:50%; padding-left:8px; padding-bottom:12px;">
                      <div style="border:1px solid #EEF2F7; border-radius:14px; padding:14px; background:#FFFFFF;">
                        <p style="margin:0 0 4px; color:#9CA3AF; font-size:11px; line-height:15px; font-weight:800; text-transform:uppercase;">Senioridade</p>
                        <p style="margin:0; color:#111827; font-size:14px; line-height:20px; font-weight:700;">{senioridade}</p>
                      </div>
                    </td>
                  </tr>
                  <tr>
                    <td class="detail-cell" width="50%" style="width:50%; padding-right:8px;">
                      <div style="border:1px solid #EEF2F7; border-radius:14px; padding:14px; background:#FFFFFF;">
                        <p style="margin:0 0 4px; color:#9CA3AF; font-size:11px; line-height:15px; font-weight:800; text-transform:uppercase;">Status</p>
                        <p style="margin:0; color:#111827; font-size:14px; line-height:20px; font-weight:700;">{status}</p>
                      </div>
                    </td>
                    <td class="detail-cell" width="50%" style="width:50%; padding-left:8px;">
                      <div style="border:1px solid #EEF2F7; border-radius:14px; padding:14px; background:#FFFFFF;">
                        <p style="margin:0 0 4px; color:#9CA3AF; font-size:11px; line-height:15px; font-weight:800; text-transform:uppercase;">Abertura</p>
                        <p style="margin:0; color:#111827; font-size:14px; line-height:20px; font-weight:700;">{data_abertura}</p>
                      </div>
                    </td>
                  </tr>
                </table>

                <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                  <tr>
                    <td align="center" style="padding-top:2px;">
                      <a class="button" href="{escape(dashboard_url, quote=True)}" style="display:inline-block; background:#6200EE; color:#FFFFFF; text-decoration:none; border-radius:12px; padding:14px 22px; font-size:14px; line-height:18px; font-weight:800;">
                        Ver no dashboard
                      </a>
                    </td>
                  </tr>
                </table>

                <p style="margin:24px 0 0; color:#6B7280; font-size:12px; line-height:19px; text-align:center;">
                  Tipo da vaga: {tipo}<br>
                  Atenciosamente, Sistema de Recrutamento
                </p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""

def notificar_avanco_vaga(db: Session, vaga: VagaModel, resumo: str) -> None:
    gestor = db.query(UserModel).filter(UserModel.id == vaga.solicitante_id).first()
    if not gestor:
        logger.warning("Notificação ignorada: gestor responsável pela vaga %s não encontrado.", vaga.id)
        return
    if not gestor.email:
        logger.warning("Notificação ignorada: gestor responsável pela vaga %s não possui e-mail.", vaga.id)
        return

    assunto = f"Avanço na vaga: {vaga.cargo}"
    corpo = (
        f"Olá{', ' + gestor.nome if gestor.nome else ''},\n\n"
        f"Houve um novo avanço na vaga \"{vaga.cargo}\".\n\n"
        "Resumo do ocorrido:\n"
        f"{resumo}\n\n"
        "Atenciosamente,\n"
        "Sistema de Recrutamento"
    )
    corpo_html = montar_email_avanco_html(vaga, gestor, resumo)

    enviar_email_notificacao(gestor.email, assunto, corpo, corpo_html)

def notificar_avanco_etapa(db: Session, vaga: VagaModel, etapa_anterior: int, nova_etapa: int) -> None:
    etapa_origem = ETAPAS_FUNIL.get(etapa_anterior, f"Etapa {etapa_anterior}")
    etapa_destino = ETAPAS_FUNIL.get(nova_etapa, f"Etapa {nova_etapa}")
    resumo = f'A etapa da vaga foi alterada de "{etapa_origem}" para "{etapa_destino}".'
    notificar_avanco_vaga(db, vaga, resumo)

def notificar_avanco_decisao(db: Session, vaga: VagaModel, status_anterior: str, novo_status: str, justificativa: Optional[str] = None) -> None:
    resumo = f'A decisão da vaga foi alterada de "{status_anterior}" para "{novo_status}".'
    if justificativa:
        resumo = f"{resumo}\nJustificativa: {justificativa}"
    notificar_avanco_vaga(db, vaga, resumo)

def gerar_senha_temporaria(tamanho: int = 14) -> str:
    tamanho = max(tamanho, 10)
    simbolos = "!@#$%&*()-_=+?"
    caracteres_obrigatorios = [
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.digits),
        secrets.choice(simbolos),
    ]
    todos = string.ascii_letters + string.digits + simbolos
    senha = caracteres_obrigatorios + [secrets.choice(todos) for _ in range(tamanho - len(caracteres_obrigatorios))]
    secrets.SystemRandom().shuffle(senha)
    return "".join(senha)

def montar_email_acesso_gestor_texto(usuario: UserModel, senha_temporaria: str, reset: bool = False) -> str:
    app_url = (os.getenv("APP_URL") or os.getenv("FRONTEND_URL") or "http://localhost:4200").rstrip("/")
    acao = "sua senha foi redefinida" if reset else "seu acesso foi criado"
    return (
        f"Olá{', ' + usuario.nome if usuario.nome else ''},\n\n"
        f"{acao.capitalize()} na plataforma FeedRH.\n\n"
        "Dados de acesso:\n"
        f"Link: {app_url}\n"
        f"E-mail: {usuario.email}\n"
        f"Senha temporária: {senha_temporaria}\n\n"
        "Por segurança, não compartilhe esta senha. Caso exista rotina de troca de senha no ambiente, "
        "altere a senha no primeiro acesso.\n\n"
        "Atenciosamente,\n"
        "Sistema de Recrutamento"
    )

def montar_email_acesso_gestor_html(usuario: UserModel, senha_temporaria: str, reset: bool = False) -> str:
    app_url = (os.getenv("APP_URL") or os.getenv("FRONTEND_URL") or "http://localhost:4200").rstrip("/")
    titulo = "Nova senha FeedRH" if reset else "Acesso ao FeedRH"
    subtitulo = "Sua senha temporária foi gerada." if reset else "Seu cadastro de gestor foi criado."
    usuario_nome = escape(usuario.nome or "Gestor")
    usuario_email = escape(usuario.email or "")
    senha = escape(senha_temporaria)
    link = escape(app_url, quote=True)

    return f"""<!doctype html>
<html lang="pt-BR">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
  </head>
  <body style="margin:0; padding:0; background:#F3E5F5; font-family:Inter, Arial, Helvetica, sans-serif; color:#1F2937;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#F3E5F5; padding:32px 12px;">
      <tr>
        <td align="center">
          <table role="presentation" width="620" cellpadding="0" cellspacing="0" style="width:620px; max-width:100%; background:#FFFFFF; border-radius:20px; overflow:hidden; box-shadow:0 18px 45px rgba(49,27,146,0.14); border:1px solid #E9D5FF;">
            <tr>
              <td style="padding:34px; background:#6200EE;">
                <p style="margin:0; color:#EDE7FF; font-size:13px; font-weight:800; letter-spacing:0.08em; text-transform:uppercase;">FeedRH</p>
                <h1 style="margin:18px 0 8px; color:#FFFFFF; font-size:30px; line-height:36px; font-weight:800;">{escape(titulo)}</h1>
                <p style="margin:0; color:#EDE7FF; font-size:15px; line-height:24px;">Olá, {usuario_nome}. {escape(subtitulo)}</p>
              </td>
            </tr>
            <tr>
              <td style="padding:30px 34px 34px;">
                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:22px;">
                  <tr>
                    <td style="background:#FAF7FF; border:1px solid #E9D5FF; border-radius:16px; padding:20px;">
                      <p style="margin:0 0 12px; color:#6200EE; font-size:12px; line-height:16px; font-weight:800; letter-spacing:0.06em; text-transform:uppercase;">Dados de acesso</p>
                      <p style="margin:0 0 8px; color:#1F2937; font-size:15px; line-height:23px;"><strong>Link:</strong> <a href="{link}" style="color:#6200EE;">{link}</a></p>
                      <p style="margin:0 0 8px; color:#1F2937; font-size:15px; line-height:23px;"><strong>E-mail:</strong> {usuario_email}</p>
                      <p style="margin:0; color:#1F2937; font-size:15px; line-height:23px;"><strong>Senha temporária:</strong> <span style="font-family:Consolas, monospace; background:#FFFFFF; border:1px solid #E9D5FF; border-radius:8px; padding:4px 8px;">{senha}</span></p>
                    </td>
                  </tr>
                </table>
                <p style="margin:0 0 22px; color:#4B5563; font-size:14px; line-height:22px;">
                  Por segurança, não compartilhe esta senha. Caso exista rotina de troca de senha no ambiente, altere a senha no primeiro acesso.
                </p>
                <div style="text-align:center;">
                  <a href="{link}" style="display:inline-block; background:#6200EE; color:#FFFFFF; text-decoration:none; border-radius:12px; padding:14px 22px; font-size:14px; line-height:18px; font-weight:800;">Acessar plataforma</a>
                </div>
                <p style="margin:24px 0 0; color:#6B7280; font-size:12px; line-height:19px; text-align:center;">Atenciosamente, Sistema de Recrutamento</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""

def enviar_email_acesso_gestor(usuario: UserModel, senha_temporaria: str, reset: bool = False) -> bool:
    if not usuario.email:
        logger.error("Gestor %s não possui e-mail para envio de acesso.", usuario.id)
        return False
    assunto = "Nova senha de acesso - FeedRH" if reset else "Acesso à plataforma FeedRH"
    corpo = montar_email_acesso_gestor_texto(usuario, senha_temporaria, reset)
    corpo_html = montar_email_acesso_gestor_html(usuario, senha_temporaria, reset)
    return enviar_email_notificacao(usuario.email, assunto, corpo, corpo_html)

def formatar_data(valor: Optional[datetime], incluir_hora: bool = False) -> str:
    if not valor:
        return ""
    formato = "%d/%m/%Y %H:%M" if incluir_hora else "%d/%m/%Y"
    return valor.strftime(formato)

def valor_relatorio(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return formatar_data(valor, incluir_hora=True)
    return str(valor)

def obter_vagas_relatorio(
    db: Session,
    gestor_id: Optional[int],
    data_inicio: Optional[str],
    data_fim: Optional[str],
    empresa: Optional[str],
    senioridade: Optional[str],
    etapa_funil: Optional[int],
    status_decisao: Optional[str],
) -> List[VagaModel]:
    query = aplicar_filtros_vagas(
        db.query(VagaModel),
        gestor_id,
        data_inicio,
        data_fim,
        empresa=empresa,
        senioridade=senioridade,
        etapa_funil=etapa_funil,
        status_decisao=status_decisao,
    )
    vagas = query.order_by(VagaModel.data_abertura.asc(), VagaModel.id.asc()).all()
    return preencher_dados_vagas(db, vagas)

def montar_linhas_relatorio(vagas: List[VagaModel]) -> List[dict]:
    linhas = []
    for vaga in vagas:
        etapa_nome = ETAPAS_FUNIL.get(vaga.etapa_funil, "Desconhecida")
        gestor_nome = getattr(vaga, "solicitante_nome", None) or f"Gestor #{vaga.solicitante_id}"
        historico = []
        for registro in getattr(vaga, "historico", []) or []:
            historico.append({
                "id": registro.id,
                "vaga_id": vaga.id,
                "vaga_cargo": vaga.cargo,
                "data_registro": registro.data_registro,
                "usuario_id": registro.usuario_id,
                "usuario_nome": registro.usuario_nome,
                "acao": registro.acao,
                "status_anterior": registro.status_anterior,
                "status_novo": registro.status_novo,
                "justificativa": registro.justificativa,
            })
        linhas.append({
            "id": vaga.id,
            "cargo": vaga.cargo,
            "empresa_destinada": vaga.empresa_destinada,
            "senioridade": vaga.senioridade,
            "tipo": vaga.tipo,
            "gestor_id": vaga.solicitante_id,
            "gestor_nome": gestor_nome,
            "gestor_email": getattr(vaga, "solicitante_email", None),
            "gestor_empresa": getattr(vaga, "solicitante_empresa", None),
            "data_abertura": vaga.data_abertura,
            "status_decisao_diretoria": vaga.status_decisao_diretoria,
            "justificativa_negativa": vaga.justificativa_negativa,
            "quantidade_congelamentos": vaga.quantidade_congelamentos or 0,
            "etapa_funil": vaga.etapa_funil,
            "etapa_nome": etapa_nome,
            "posicao_fila_rh": vaga.posicao_fila_rh,
            "data_finalizacao": vaga.data_finalizacao,
            "resumo_requisitos": vaga.resumo_requisitos,
            "requisitos_obrigatorios": vaga.requisitos_obrigatorios,
            "profissional_substituido": vaga.profissional_substituido,
            "justificativa_substituicao": vaga.justificativa_substituicao,
            "historico": historico,
        })
    return linhas

def contar_por(linhas: List[dict], campo: str) -> dict:
    totais = {}
    for linha in linhas:
        chave = linha.get(campo) or "Não informado"
        totais[chave] = totais.get(chave, 0) + 1
    return totais

def gerar_resumo_relatorio(linhas: List[dict]) -> dict:
    return {
        "total_vagas": len(linhas),
        "por_status": contar_por(linhas, "status_decisao_diretoria"),
        "por_etapa": contar_por(linhas, "etapa_nome"),
        "por_gestor": contar_por(linhas, "gestor_nome"),
        "por_empresa": contar_por(linhas, "empresa_destinada"),
    }

def descrever_filtros_relatorio(
    db: Session,
    gestor_id: Optional[int],
    data_inicio: Optional[str],
    data_fim: Optional[str],
    empresa: Optional[str],
    senioridade: Optional[str],
    etapa_funil: Optional[int],
    status_decisao: Optional[str],
) -> List[str]:
    filtros = []
    if gestor_id is not None:
        gestor = db.query(UserModel).filter(UserModel.id == gestor_id).first()
        filtros.append(f"Gestor: {gestor.nome if gestor else 'ID ' + str(gestor_id)}")
    if empresa:
        filtros.append(f"Empresa: {empresa}")
    if senioridade:
        filtros.append(f"Senioridade: {senioridade}")
    if etapa_funil is not None:
        filtros.append(f"Etapa: {ETAPAS_FUNIL.get(etapa_funil, 'Etapa ' + str(etapa_funil))}")
    if status_decisao:
        filtros.append(f"Status: {status_decisao}")
    if data_inicio:
        filtros.append(f"Data inicial: {data_inicio}")
    if data_fim:
        filtros.append(f"Data final: {data_fim}")
    return filtros or ["Nenhum filtro aplicado"]

def paragrafo_pdf(texto, estilo):
    texto_seguro = escape(valor_relatorio(texto)).replace("\n", "<br/>")
    return Paragraph(texto_seguro or "-", estilo)

def gerar_pdf_relatorio(linhas: List[dict], resumo: dict, filtros: List[str]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("FeedRHTitulo", parent=styles["Title"], fontSize=18, leading=22, textColor=colors.HexColor("#311B92"))
    subtitulo_style = ParagraphStyle("FeedRHSubtitulo", parent=styles["Heading2"], fontSize=12, leading=15, textColor=colors.HexColor("#4B5563"))
    normal_style = ParagraphStyle("FeedRHNormal", parent=styles["BodyText"], fontSize=8, leading=10)
    small_style = ParagraphStyle("FeedRHSmall", parent=styles["BodyText"], fontSize=7, leading=9)
    header_style = ParagraphStyle("FeedRHHeader", parent=styles["BodyText"], fontSize=7, leading=9, textColor=colors.white)

    elementos = [
        Paragraph("Relatório de Vagas - FeedRH", titulo_style),
        Paragraph(f"Gerado em {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC", normal_style),
        Spacer(1, 0.25 * cm),
        Paragraph("Filtros aplicados", subtitulo_style),
        Paragraph("<br/>".join(escape(filtro) for filtro in filtros), normal_style),
        Spacer(1, 0.25 * cm),
        Paragraph("Resumo", subtitulo_style),
    ]

    resumo_linhas = [["Métrica", "Valor"]]
    resumo_linhas.append(["Total de vagas", str(resumo["total_vagas"])])
    for titulo, chave in [
        ("Total por status", "por_status"),
        ("Total por etapa", "por_etapa"),
        ("Total por gestor", "por_gestor"),
        ("Total por empresa", "por_empresa"),
    ]:
        for nome, total in resumo[chave].items():
            resumo_linhas.append([titulo, f"{nome}: {total}"])
    tabela_resumo = Table(resumo_linhas, colWidths=[5 * cm, 12 * cm])
    tabela_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6200EE")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    elementos.extend([tabela_resumo, Spacer(1, 0.35 * cm), Paragraph("Tabela principal de vagas", subtitulo_style)])

    cabecalho = ["ID", "Cargo", "Empresa", "Gestor", "Abertura", "Status", "Etapa", "Fila", "Cong."]
    tabela_vagas = [[paragrafo_pdf(item, header_style) for item in cabecalho]]
    for linha in linhas:
        tabela_vagas.append([
            paragrafo_pdf(linha["id"], small_style),
            paragrafo_pdf(linha["cargo"], small_style),
            paragrafo_pdf(linha["empresa_destinada"], small_style),
            paragrafo_pdf(linha["gestor_nome"], small_style),
            paragrafo_pdf(formatar_data(linha["data_abertura"]), small_style),
            paragrafo_pdf(linha["status_decisao_diretoria"], small_style),
            paragrafo_pdf(linha["etapa_nome"], small_style),
            paragrafo_pdf(linha["posicao_fila_rh"] or "", small_style),
            paragrafo_pdf(linha["quantidade_congelamentos"], small_style),
        ])
    tabela_principal = Table(
        tabela_vagas,
        repeatRows=1,
        colWidths=[1.0 * cm, 4.5 * cm, 3.5 * cm, 3.8 * cm, 2.1 * cm, 2.2 * cm, 3.4 * cm, 1.2 * cm, 1.2 * cm],
    )
    tabela_principal.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6200EE")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAF7FF")]),
    ]))
    elementos.append(tabela_principal)

    if linhas:
        elementos.extend([PageBreak(), Paragraph("Detalhes e histórico por vaga", subtitulo_style)])
    for linha in linhas:
        elementos.append(Spacer(1, 0.18 * cm))
        elementos.append(Paragraph(f"Vaga #{linha['id']} - {escape(linha['cargo'] or '')}", subtitulo_style))
        detalhes = [
            ["Campo", "Valor"],
            ["Tipo", linha["tipo"]],
            ["Gestor/solicitante", f"{linha['gestor_nome']} <{linha.get('gestor_email') or ''}>"],
            ["Empresa do gestor", linha.get("gestor_empresa") or ""],
            ["Data de finalização", formatar_data(linha.get("data_finalizacao"))],
            ["Justificativa de negativa", linha.get("justificativa_negativa") or ""],
            ["Resumo dos requisitos", linha.get("resumo_requisitos") or ""],
            ["Requisitos obrigatórios", linha.get("requisitos_obrigatorios") or ""],
            ["Profissional substituído", linha.get("profissional_substituido") or ""],
            ["Justificativa da substituição", linha.get("justificativa_substituicao") or ""],
        ]
        tabela_detalhes = Table(
            [[paragrafo_pdf(celula, small_style) for celula in row] for row in detalhes],
            colWidths=[5 * cm, 20 * cm],
        )
        tabela_detalhes.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#311B92")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.append(tabela_detalhes)

        if linha["historico"]:
            historico_pdf = [["Data", "Usuário", "Ação", "Anterior", "Novo", "Justificativa"]]
            for registro in linha["historico"]:
                historico_pdf.append([
                    formatar_data(registro["data_registro"], incluir_hora=True),
                    registro["usuario_nome"],
                    registro["acao"],
                    registro["status_anterior"] or "",
                    registro["status_novo"] or "",
                    registro["justificativa"] or "",
                ])
            tabela_historico = Table(
                [[paragrafo_pdf(celula, small_style) for celula in row] for row in historico_pdf],
                colWidths=[3 * cm, 4 * cm, 4 * cm, 3 * cm, 3 * cm, 8 * cm],
            )
            tabela_historico.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6200EE")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            elementos.extend([Spacer(1, 0.15 * cm), tabela_historico])

    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()

def aplicar_estilo_cabecalho(ws):
    fill = PatternFill("solid", fgColor="6200EE")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

def ajustar_larguras(ws):
    for coluna in ws.columns:
        maior = 0
        letra = get_column_letter(coluna[0].column)
        for cell in coluna:
            maior = max(maior, len(str(cell.value or "")))
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 60)

def gerar_excel_relatorio(linhas: List[dict], resumo: dict) -> bytes:
    workbook = Workbook()
    ws_vagas = workbook.active
    ws_vagas.title = "Vagas"
    colunas_vagas = [
        ("id", "ID da vaga"),
        ("cargo", "Cargo"),
        ("empresa_destinada", "Empresa destinada"),
        ("senioridade", "Senioridade"),
        ("tipo", "Tipo da vaga"),
        ("gestor_nome", "Gestor/solicitante"),
        ("gestor_email", "E-mail do gestor"),
        ("gestor_empresa", "Empresa do gestor"),
        ("data_abertura", "Data de abertura"),
        ("status_decisao_diretoria", "Status da decisão"),
        ("justificativa_negativa", "Justificativa de negativa"),
        ("quantidade_congelamentos", "Quantidade de congelamentos"),
        ("etapa_funil", "Etapa atual do funil"),
        ("etapa_nome", "Nome da etapa"),
        ("posicao_fila_rh", "Posição na fila RH"),
        ("data_finalizacao", "Data de finalização"),
        ("resumo_requisitos", "Resumo dos requisitos"),
        ("requisitos_obrigatorios", "Requisitos obrigatórios"),
        ("profissional_substituido", "Profissional substituído"),
        ("justificativa_substituicao", "Justificativa da substituição"),
    ]
    ws_vagas.append([titulo for _, titulo in colunas_vagas])
    for linha in linhas:
        ws_vagas.append([valor_relatorio(linha.get(campo)) for campo, _ in colunas_vagas])
    aplicar_estilo_cabecalho(ws_vagas)
    ajustar_larguras(ws_vagas)

    ws_resumo = workbook.create_sheet("Resumo")
    ws_resumo.append(["Grupo", "Item", "Total"])
    ws_resumo.append(["Total", "Vagas", resumo["total_vagas"]])
    for grupo, chave in [
        ("Status", "por_status"),
        ("Etapa", "por_etapa"),
        ("Gestor", "por_gestor"),
        ("Empresa", "por_empresa"),
    ]:
        for item, total in resumo[chave].items():
            ws_resumo.append([grupo, item, total])
    aplicar_estilo_cabecalho(ws_resumo)
    ajustar_larguras(ws_resumo)

    ws_historico = workbook.create_sheet("Histórico")
    ws_historico.append([
        "ID da vaga",
        "Cargo",
        "Data do registro",
        "Usuário responsável",
        "Ação realizada",
        "Status anterior",
        "Status novo",
        "Justificativa",
    ])
    for linha in linhas:
        for registro in linha["historico"]:
            ws_historico.append([
                registro["vaga_id"],
                registro["vaga_cargo"],
                valor_relatorio(registro["data_registro"]),
                registro["usuario_nome"],
                registro["acao"],
                registro["status_anterior"] or "",
                registro["status_novo"] or "",
                registro["justificativa"] or "",
            ])
    aplicar_estilo_cabecalho(ws_historico)
    ajustar_larguras(ws_historico)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# Seed default data on an empty database
@app.on_event("startup")
def startup_event():
    initialize_database()
    db = SessionLocal()
    try:
        if not db.query(EmpresaModel).first():
            empresas_padrao = ["Elevare", "Ora Empresas", "Mercado do Provedor", "Mercado do Construtor", "Outra"]
            for empresa in empresas_padrao:
                criar_empresa_se_nao_existir(db, empresa)

        if not db.query(UserModel).first():
            rh_user = UserModel(
                nome="Fernanda Silva",
                email="rh@feedrh.com",
                empresa="Elevare",
                perfil="RH",
                senha_hash=hash_senha("rh@123")
            )
            gestor_user = UserModel(
                nome="Carlos Souza",
                email="gestor@feedrh.com",
                empresa="Ora Empresas",
                perfil="GESTOR",
                senha_hash=hash_senha("gestor@123")
            )
            db.add(rh_user)
            db.add(gestor_user)
            db.commit()
    finally:
        db.close()

# --- Routes ---

@app.post("/auth/login", response_model=LoginResponse)
def login(credentials: LoginRequest, db: Session = Depends(get_db)):
    email = credentials.email.strip().lower()
    user = db.query(UserModel).filter(UserModel.email.ilike(email)).first()
    if not user or user.senha_hash != hash_senha(credentials.senha):
        raise HTTPException(status_code=401, detail="E-mail ou senha inválidos.")
    return LoginResponse(
        id=user.id,
        nome=user.nome,
        email=user.email,
        empresa=user.empresa,
        perfil=user.perfil
    )

@app.post("/users", response_model=UserCreateResponse)
def create_user(user: UserCreate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    nome_normalizado, email_normalizado, empresa_normalizada, perfil = validar_dados_usuario(
        db,
        user.nome,
        user.email,
        user.empresa,
        user.perfil
    )
    if not user.senha or len(user.senha) < 6:
        raise HTTPException(status_code=400, detail="A senha deve ter no mínimo 6 caracteres")

    agora = datetime.utcnow()
    db_user = UserModel(
        nome=nome_normalizado,
        email=email_normalizado,
        empresa=empresa_normalizada,
        perfil=perfil,
        senha_hash=hash_senha(user.senha),
        must_change_password=False,
        created_at=agora,
        updated_at=agora,
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)

    email_enviado = None
    if perfil == "GESTOR":
        email_enviado = enviar_email_acesso_gestor(db_user, user.senha, reset=False)
        if not email_enviado:
            logger.error("Gestor %s criado, mas o e-mail de acesso não foi enviado.", db_user.id)

    return UserCreateResponse(
        id=db_user.id,
        nome=db_user.nome,
        email=db_user.email,
        empresa=db_user.empresa,
        perfil=db_user.perfil,
        must_change_password=bool(db_user.must_change_password),
        created_at=db_user.created_at,
        updated_at=db_user.updated_at,
        ultimo_reset_senha=db_user.ultimo_reset_senha,
        message="Usuário criado com sucesso",
        email_enviado=email_enviado,
    )

@app.get("/users", response_model=List[UserResponse])
def get_users(db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    return db.query(UserModel).all()

@app.patch("/users/{user_id}", response_model=UserResponse)
def update_user(user_id: int, update: UserUpdate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    nome_normalizado, email_normalizado, empresa_normalizada, perfil = validar_dados_usuario(
        db,
        update.nome,
        update.email,
        update.empresa,
        update.perfil,
        user_id_ignorar=user_id
    )

    if db_user.perfil == "RH" and perfil != "RH":
        total_rh = db.query(UserModel).filter(UserModel.perfil == "RH").count()
        if total_rh <= 1:
            raise HTTPException(status_code=400, detail="Não é possível remover o último usuário RH")
        if db_user.id == current_user.id:
            raise HTTPException(status_code=400, detail="Não é possível remover seu próprio acesso de RH")

    db_user.nome = nome_normalizado
    db_user.email = email_normalizado
    db_user.empresa = empresa_normalizada
    db_user.perfil = perfil
    db_user.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/users/{user_id}/reset-password", response_model=ResetPasswordResponse)
def reset_password_gestor(user_id: int, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if db_user.perfil != "GESTOR":
        raise HTTPException(status_code=400, detail="Apenas senhas de gestores podem ser resetadas nesta versão")

    senha_temporaria = gerar_senha_temporaria()
    agora = datetime.utcnow()
    db_user.senha_hash = hash_senha(senha_temporaria)
    db_user.ultimo_reset_senha = agora
    db_user.updated_at = agora
    db.commit()
    db.refresh(db_user)

    email_enviado = enviar_email_acesso_gestor(db_user, senha_temporaria, reset=True)
    logger.info(
        "Reset de senha solicitado pelo RH %s para gestor %s. E-mail enviado: %s",
        current_user.id,
        db_user.id,
        email_enviado,
    )
    return ResetPasswordResponse(
        message="Senha resetada com sucesso",
        email_enviado=email_enviado,
    )

@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_user = db.query(UserModel).filter(UserModel.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Não é possível apagar sua própria conta")
    if db_user.perfil == "RH":
        total_rh = db.query(UserModel).filter(UserModel.perfil == "RH").count()
        if total_rh <= 1:
            raise HTTPException(status_code=400, detail="Não é possível apagar o último usuário RH")

    db.delete(db_user)
    db.commit()
    return {"detail": "Usuário apagado com sucesso"}

@app.get("/empresas", response_model=List[EmpresaResponse])
def get_empresas(db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    return db.query(EmpresaModel).order_by(EmpresaModel.nome.asc()).all()

@app.post("/empresas", response_model=EmpresaResponse)
def create_empresa(empresa: EmpresaCreate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    nome_normalizado = validar_nome_empresa(empresa.nome)
    if buscar_empresa_por_nome(db, nome_normalizado):
        raise HTTPException(status_code=400, detail="Empresa já cadastrada")
    db_empresa = EmpresaModel(nome=nome_normalizado)
    db.add(db_empresa)
    db.commit()
    db.refresh(db_empresa)
    return db_empresa

@app.patch("/empresas/{empresa_id}", response_model=EmpresaResponse)
def update_empresa(empresa_id: int, update: EmpresaUpdate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_empresa = db.query(EmpresaModel).filter(EmpresaModel.id == empresa_id).first()
    if not db_empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    nome_normalizado = validar_nome_empresa(update.nome)
    empresa_existente = buscar_empresa_por_nome(db, nome_normalizado)
    if empresa_existente and empresa_existente.id != empresa_id:
        raise HTTPException(status_code=400, detail="Empresa já cadastrada")

    nome_anterior = db_empresa.nome
    db_empresa.nome = nome_normalizado
    db.query(UserModel).filter(UserModel.empresa == nome_anterior).update(
        {UserModel.empresa: nome_normalizado},
        synchronize_session=False
    )
    db.query(VagaModel).filter(VagaModel.empresa_destinada == nome_anterior).update(
        {VagaModel.empresa_destinada: nome_normalizado},
        synchronize_session=False
    )
    db.commit()
    db.refresh(db_empresa)
    return db_empresa

@app.delete("/empresas/{empresa_id}")
def delete_empresa(empresa_id: int, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_empresa = db.query(EmpresaModel).filter(EmpresaModel.id == empresa_id).first()
    if not db_empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")

    usuarios_vinculados = db.query(UserModel).filter(UserModel.empresa == db_empresa.nome).count()
    vagas_vinculadas = db.query(VagaModel).filter(VagaModel.empresa_destinada == db_empresa.nome).count()
    if usuarios_vinculados or vagas_vinculadas:
        raise HTTPException(
            status_code=400,
            detail="Empresa possui usuários ou vagas vinculadas. Edite o nome em vez de apagar."
        )

    db.delete(db_empresa)
    db.commit()
    return {"detail": "Empresa apagada com sucesso"}

@app.post("/vagas", response_model=VagaResponse)
def create_vaga(vaga: VagaCreate, db: Session = Depends(get_db), current_user: UserModel = Depends(get_current_user)):
    empresa_destinada = normalizar_nome_empresa(vaga.empresa_destinada)
    if not buscar_empresa_por_nome(db, empresa_destinada):
        raise HTTPException(status_code=400, detail="Empresa destinada inválida")
    db_vaga = VagaModel(
        cargo=vaga.cargo,
        empresa_destinada=empresa_destinada,
        senioridade=vaga.senioridade,
        resumo_requisitos=vaga.resumo_requisitos.strip(),
        requisitos_obrigatorios=vaga.requisitos_obrigatorios.strip(),
        tipo=vaga.tipo,
        profissional_substituido=vaga.profissional_substituido,
        justificativa_substituicao=vaga.justificativa_substituicao,
        solicitante_id=current_user.id,
        status_decisao_diretoria="Pendente",
        quantidade_congelamentos=0,
        etapa_funil=1
    )
    db.add(db_vaga)
    db.commit()
    db.refresh(db_vaga)
    return preencher_dados_vagas(db, [db_vaga])[0]

@app.get("/vagas", response_model=List[VagaResponse])
def get_vagas(
    gestor_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    empresa: Optional[str] = Query(None),
    senioridade: Optional[str] = Query(None),
    etapa_funil: Optional[int] = Query(None),
    status_decisao: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user)
):
    if current_user.perfil == "RH":
        query = db.query(VagaModel)
        query = aplicar_filtros_vagas(
            query,
            gestor_id,
            data_inicio,
            data_fim,
            empresa=empresa,
            senioridade=senioridade,
            etapa_funil=etapa_funil,
            status_decisao=status_decisao,
        )
    else:
        query = db.query(VagaModel).filter(VagaModel.solicitante_id == current_user.id)
        query = aplicar_filtros_vagas(
            query,
            None,
            data_inicio,
            data_fim,
            empresa=empresa,
            senioridade=senioridade,
            etapa_funil=etapa_funil,
            status_decisao=status_decisao,
        )
    vagas = query.order_by(VagaModel.data_abertura.asc(), VagaModel.id.asc()).all()
    return preencher_dados_vagas(db, vagas)

@app.patch("/vagas/{vaga_id}/decisao-diretoria", response_model=VagaResponse)
def update_decisao_diretoria(vaga_id: int, update: DecisaoDiretoriaUpdate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_vaga = db.query(VagaModel).filter(VagaModel.id == vaga_id).first()
    if not db_vaga:
        raise HTTPException(status_code=404, detail="Vaga não encontrada")
    if update.status not in ["Pendente", "Aprovada", "Congelada", "Negada"]:
        raise HTTPException(status_code=400, detail="Status de decisão inválido")
    justificativa_negativa = (update.justificativa_negativa or "").strip()
    if update.status == "Negada" and not justificativa_negativa:
        raise HTTPException(status_code=400, detail="Informe a justificativa para negar a vaga")

    status_anterior = db_vaga.status_decisao_diretoria
    db_vaga.status_decisao_diretoria = update.status
    if update.status == "Negada":
        db_vaga.justificativa_negativa = justificativa_negativa
    if update.status == "Congelada" and status_anterior != "Congelada":
        db_vaga.quantidade_congelamentos = (db_vaga.quantidade_congelamentos or 0) + 1
    houve_avanco = status_anterior != update.status or update.status == "Negada"
    if houve_avanco:
        db.add(VagaHistoricoModel(
            vaga_id=db_vaga.id,
            usuario_id=current_user.id,
            usuario_nome=current_user.nome,
            acao="Decisão do RH",
            status_anterior=status_anterior,
            status_novo=update.status,
            justificativa=justificativa_negativa or None,
        ))
    db.commit()
    db.refresh(db_vaga)
    if houve_avanco:
        notificar_avanco_decisao(db, db_vaga, status_anterior, update.status, justificativa_negativa or None)
    return preencher_dados_vagas(db, [db_vaga])[0]

@app.patch("/vagas/{vaga_id}/etapa-funil", response_model=VagaResponse)
def update_etapa_funil(vaga_id: int, update: EtapaFunilUpdate, db: Session = Depends(get_db), current_user: UserModel = Depends(require_rh)):
    db_vaga = db.query(VagaModel).filter(VagaModel.id == vaga_id).first()
    if not db_vaga:
        raise HTTPException(status_code=404, detail="Vaga não encontrada")
    if update.etapa < 1 or update.etapa > 9:
        raise HTTPException(status_code=400, detail="Etapa deve ser de 1 a 9")
    etapa_anterior = db_vaga.etapa_funil or 1
    db_vaga.etapa_funil = update.etapa
    if update.etapa == 9:
        db_vaga.data_finalizacao = datetime.utcnow()
    else:
        db_vaga.data_finalizacao = None
    db.commit()
    db.refresh(db_vaga)
    if update.etapa != etapa_anterior:
        notificar_avanco_etapa(db, db_vaga, etapa_anterior, update.etapa)
    return preencher_dados_vagas(db, [db_vaga])[0]

@app.get("/vagas/relatorio")
def get_relatorio(
    gestor_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    empresa: Optional[str] = Query(None),
    senioridade: Optional[str] = Query(None),
    etapa_funil: Optional[int] = Query(None),
    status_decisao: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(require_rh)
):
    vagas = obter_vagas_relatorio(db, gestor_id, data_inicio, data_fim, empresa, senioridade, etapa_funil, status_decisao)
    linhas = montar_linhas_relatorio(vagas)
    resumo = gerar_resumo_relatorio(linhas)
    finalizadas_no_mes = 0
    now = datetime.utcnow()
    for linha in linhas:
        data_finalizacao = linha.get("data_finalizacao")
        if linha.get("etapa_funil") == 9 and data_finalizacao:
            if data_finalizacao.year == now.year and data_finalizacao.month == now.month:
                finalizadas_no_mes += 1
    return {
        "total_abertas": resumo["total_vagas"],
        "total_aprovadas": resumo["por_status"].get("Aprovada", 0),
        "total_congeladas": resumo["por_status"].get("Congelada", 0),
        "total_negadas": resumo["por_status"].get("Negada", 0),
        "total_finalizadas_no_mes": finalizadas_no_mes,
        "agrupado_por_empresa": resumo["por_empresa"],
        "agrupado_por_senioridade": contar_por(linhas, "senioridade"),
        "agrupado_por_etapa": resumo["por_etapa"],
        "agrupado_por_gestor": resumo["por_gestor"],
        "vagas": linhas,
    }

@app.get("/vagas/relatorio/pdf")
def exportar_relatorio_pdf(
    gestor_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    empresa: Optional[str] = Query(None),
    senioridade: Optional[str] = Query(None),
    etapa_funil: Optional[int] = Query(None),
    status_decisao: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(require_rh)
):
    vagas = obter_vagas_relatorio(db, gestor_id, data_inicio, data_fim, empresa, senioridade, etapa_funil, status_decisao)
    linhas = montar_linhas_relatorio(vagas)
    resumo = gerar_resumo_relatorio(linhas)
    filtros = descrever_filtros_relatorio(db, gestor_id, data_inicio, data_fim, empresa, senioridade, etapa_funil, status_decisao)
    arquivo = gerar_pdf_relatorio(linhas, resumo, filtros)
    nome_arquivo = f"feedrh-relatorio-vagas-{datetime.utcnow().strftime('%Y-%m-%d')}.pdf"
    return StreamingResponse(
        BytesIO(arquivo),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )

@app.get("/vagas/relatorio/excel")
def exportar_relatorio_excel(
    gestor_id: Optional[int] = Query(None),
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    empresa: Optional[str] = Query(None),
    senioridade: Optional[str] = Query(None),
    etapa_funil: Optional[int] = Query(None),
    status_decisao: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(require_rh)
):
    vagas = obter_vagas_relatorio(db, gestor_id, data_inicio, data_fim, empresa, senioridade, etapa_funil, status_decisao)
    linhas = montar_linhas_relatorio(vagas)
    resumo = gerar_resumo_relatorio(linhas)
    arquivo = gerar_excel_relatorio(linhas, resumo)
    nome_arquivo = f"feedrh-relatorio-vagas-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        BytesIO(arquivo),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
